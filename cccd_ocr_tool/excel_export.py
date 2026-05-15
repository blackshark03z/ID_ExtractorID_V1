import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

def create_excel_report(data_list, output_path):
    """
    Exports data to Excel with formatting.
    data_list: List of dictionaries containing the parsed fields + FolderName.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CCCD Data"
    
    headers = [
        "HoTen", "CCCD", "NgaySinh", "GioiTinh", 
        "DiaChi", "NoiCap", "NgayCap", "NgayHetHan", "GhiChu"
    ]
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        
    # Write data
    for row_num, data in enumerate(data_list, 2):
        for col_num, header in enumerate(headers, 1):
            val = data.get(header, "")
            # Ensure numbers don't lose leading zeros by treating as string
            if header == "CCCD" and val:
                val = str(val)
            ws.cell(row=row_num, column=col_num, value=val)
            
    # Freeze top row
    ws.freeze_panes = "A2"
    
    # Auto width columns
    for col_num, header in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        column = ws[column_letter]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        # Limit max width to avoid ridiculously wide columns for long addresses/notes
        if adjusted_width > 50:
            adjusted_width = 50
        ws.column_dimensions[column_letter].width = adjusted_width
        
    wb.save(output_path)
